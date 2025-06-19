import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Tuple
import logging
import pprint
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Set your filenames here
XML_FILE = 'hierarchy.xml'
GA_FILE = 'ga_hierarchy.txt'

def safe_get_text(elem):
    if elem is not None and elem.text is not None:
        return elem.text.strip()
    return None

def build_hierarchy_tree(hier_elem) -> Dict[str, Any]:
    def build_node(dim_elem):
        name = safe_get_text(dim_elem.find('./attribute[@name="rpas_name"]/value'))
        aggs = safe_get_text(dim_elem.find('./attribute[@name="aggs"]/value'))
        userdim_val = safe_get_text(dim_elem.find('./attribute[@name="userdim"]/value'))
        userdim = userdim_val is not None and userdim_val.lower() == 'true'
        virthier = safe_get_text(dim_elem.find('./attribute[@name="virthier"]/value'))
        children = [build_node(child) for child in dim_elem.findall('./data_model[@type="DimClass"]')]
        return {
            'name': name,
            'aggs': aggs,
            'userdim': userdim,
            'virthier': virthier,
            'children': children
        }
    root_dim = hier_elem.find('./data_model[@type="DimClass"]')
    if root_dim is not None:
        return build_node(root_dim)
    return None

def get_all_paths(tree: Dict[str, Any], path=None) -> List[List[str]]:
    if path is None:
        path = []
    if tree is None or tree['name'] is None:
        return []
    current_path = path + [tree['name']]
    if not tree['children']:
        return [current_path]
    paths = []
    for child in tree['children']:
        paths.extend(get_all_paths(child, current_path))
    return paths

def extract_hierarchy_trees(xml_file: str) -> Dict[str, Any]:
    tree = ET.parse(xml_file)
    root = tree.getroot()
    hier_trees = {}
    for hier in root.findall('.//data_model[@type="HierClass"]'):
        hier_name = hier.find('./attribute[@name="rpas_name"]/value')
        if hier_name is not None and hier_name.text.strip() in ['CLND', 'PROD', 'LOC']:
            hier_trees[hier_name.text.strip().lower()] = build_hierarchy_tree(hier)
    return hier_trees

def read_ga_hierarchy_from_file(file_path: str = 'ga_hierarchy.txt') -> Dict[str, List[str]]:
    hierarchies: Dict[str, List[str]] = {}
    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line and ':' in line:
                hier_name, dimensions_str = line.split(':', 1)
                hier_name = hier_name.strip().lower()
                dimensions = [dim.strip() for dim in dimensions_str.split(',')]
                hierarchies[hier_name] = dimensions
    return hierarchies

def ordered_alignment_table(ga_list: List[str], customer_path: List[str]) -> Tuple[List[Tuple[str, str, str]], List[str], List[str]]:
    """
    Aligns GA and customer path in order, matching each GA dimension to the next available customer dimension.
    Returns:
        - List of (GA, Customer, Match) tuples for the table
        - List of unmatched GA dimensions
        - List of unmatched customer dimensions
    """
    table = []
    ga_idx = 0
    cust_idx = 0
    matched_cust_indices = set()
    unmatched_ga = []
    unmatched_cust = []
    # For each GA dimension, find the next matching customer dimension
    while ga_idx < len(ga_list):
        ga_dim = ga_list[ga_idx]
        found = False
        for i in range(cust_idx, len(customer_path)):
            if customer_path[i].lower() == ga_dim.lower():
                table.append((ga_dim, customer_path[i], '✓'))
                matched_cust_indices.add(i)
                cust_idx = i + 1
                found = True
                break
        if not found:
            table.append((ga_dim, '', '✗'))
            unmatched_ga.append(ga_dim)
        ga_idx += 1
    # Any customer dimensions not matched are extra
    for i, dim in enumerate(customer_path):
        if i not in matched_cust_indices:
            unmatched_cust.append(dim)
    return table, unmatched_ga, unmatched_cust

def find_main_spine(paths: List[List[str]], ga_list: List[str]) -> Tuple[List[str], int]:
    """
    Find the path with the most matches to the GA list (main spine), using ordered matching.
    Returns (main_spine, match_count)
    If no matches, returns the first path as the main spine (if any).
    """
    best_path = []
    best_score = 0
    for path in paths:
        table, _, _ = ordered_alignment_table(ga_list, path)
        score = sum(1 for _, _, match in table if match == '✓')
        if score > best_score:
            best_score = score
            best_path = path
    # If no matches, return the first path if available
    if not best_path and paths:
        best_path = paths[0]
        best_score = 0
    return best_path, best_score

def extract_user_defined_and_virthier(tree: Dict[str, Any], user_dims=None, virthiers=None):
    if user_dims is None:
        user_dims = []
    if virthiers is None:
        virthiers = set()
    if tree is None:
        return user_dims, virthiers
    if tree.get('userdim'):
        user_dims.append(tree['name'])
    if tree.get('virthier') == '':
        virthiers.add(tree['name'])
    for child in tree.get('children', []):
        extract_user_defined_and_virthier(child, user_dims, virthiers)
    return user_dims, virthiers

def extract_labeled_intersections(xml_file: str) -> dict:
    """
    Extract all labeled intersections from the XML file.
    Returns a dictionary: {intersection_name: intersection_definition}
    """
    tree = ET.parse(xml_file)
    root = tree.getroot()
    labeled_intersections = {}
    for intx in root.findall('.//data_model[@type="LabeledIntx"]'):
        name_elem = intx.find('./attribute[@name="name"]/value')
        def_elem = intx.find('./attribute[@name="LabeledIntxDefinition"]/value')
        if name_elem is not None and def_elem is not None:
            name = name_elem.text.strip()
            definition = def_elem.text.strip()
            labeled_intersections[name] = definition
    return labeled_intersections

def get_subtree_for_path(tree: dict, path: list) -> dict:
    """
    Given a tree and a path (list of node names), return the subtree rooted at the end of the path.
    """
    if not path or tree is None:
        return None
    node = tree
    for name in path:
        if node['name'] == name:
            if name == path[-1]:
                return node
            # Find the next child in the path
            found = False
            for child in node.get('children', []):
                if child['name'] == path[path.index(name)+1]:
                    node = child
                    found = True
                    break
            if not found:
                return None
        else:
            return None
    return node

def write_excel_output(hierarchies_data, labeled_intersections, filename='hierarchy_analysis.xlsx'):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for hier_name, data in hierarchies_data.items():
        ws = wb.create_sheet(title=hier_name.upper())
        row = 1
        # 1. Level Comparison Table
        ws.cell(row=row, column=1, value='Level').font = Font(bold=True)
        ws.cell(row=row, column=2, value='GA').font = Font(bold=True)
        ws.cell(row=row, column=3, value='Customer').font = Font(bold=True)
        ws.cell(row=row, column=4, value='Matches').font = Font(bold=True)
        row += 1
        for i, (ga, cust, match) in enumerate(data['level_comparison'], 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=ga)
            ws.cell(row=row, column=3, value=cust)
            ws.cell(row=row, column=4, value=match)
            row += 1
        row += 1
        # 2. Summary Table
        ws.cell(row=row, column=1, value='Hierarchy').font = Font(bold=True)
        ws.cell(row=row, column=2, value='User Defined Dimensions').font = Font(bold=True)
        ws.cell(row=row, column=3, value='Match Score').font = Font(bold=True)
        ws.cell(row=row, column=4, value='Virtual Hierarchy').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value=hier_name.upper())
        ws.cell(row=row, column=2, value=', '.join(data['user_dims']) if data['user_dims'] else 'None')
        ws.cell(row=row, column=3, value=data['match_score'])
        ws.cell(row=row, column=4, value=', '.join(data['virthiers']) if data['virthiers'] else 'None')
        row += 2
        # 3. Position Table
        ws.cell(row=row, column=1, value='Position').font = Font(bold=True)
        ws.cell(row=row, column=2, value='Dimension Name').font = Font(bold=True)
        ws.cell(row=row, column=3, value='In GA').font = Font(bold=True)
        row += 1
        for i, (dim, in_ga) in enumerate(data['position_table'], 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=dim)
            ws.cell(row=row, column=3, value='Yes' if in_ga else 'No')
            row += 1
        row += 2
        # 4. Path Table
        ws.cell(row=row, column=1, value='Path Type').font = Font(bold=True)
        ws.cell(row=row, column=2, value='Complete Path').font = Font(bold=True)
        ws.cell(row=row, column=3, value='Length').font = Font(bold=True)
        ws.cell(row=row, column=4, value='Is Main Spine').font = Font(bold=True)
        row += 1
        for path_type, path, length, is_main in data['path_table']:
            ws.cell(row=row, column=1, value=path_type)
            ws.cell(row=row, column=2, value=path)
            ws.cell(row=row, column=3, value=length)
            ws.cell(row=row, column=4, value='Yes' if is_main else 'No')
            row += 1
    # 5. Labeled Intersections Table (separate sheet)
    ws_intx = wb.create_sheet(title='LabeledIntersections')
    ws_intx.cell(row=1, column=1, value='Name').font = Font(bold=True)
    ws_intx.cell(row=1, column=2, value='Definition').font = Font(bold=True)
    for i, (k, v) in enumerate(labeled_intersections.items(), 2):
        ws_intx.cell(row=i, column=1, value=k)
        ws_intx.cell(row=i, column=2, value=v)
    wb.save(filename)
    wb.close()

def output_hierarchy_analysis(xml_file: str, ga_hierarchies: Dict[str, List[str]]):
    hier_trees = extract_hierarchy_trees(xml_file)
    labeled_intx = extract_labeled_intersections(xml_file)
    hierarchies_data = {}
    for hier_name, tree in hier_trees.items():
        print(f"\nHierarchy: {hier_name.upper()}")
        all_paths = get_all_paths(tree)
        print("All valid paths (including alternates):")
        for i, path in enumerate(all_paths):
            print(f"  Path {i+1}: {' -> '.join(path)}")
        ga_list = ga_hierarchies.get(hier_name, [])
        main_spine, match_score = find_main_spine(all_paths, ga_list)
        if main_spine:
            print(f"\nMain spine (most matches with GA): {' -> '.join(main_spine)} (Matches: {match_score})")
            subtree = get_subtree_for_path(tree, main_spine)
            print("Main spine subtree (full data structure):")
            pprint.pprint(subtree, indent=2, width=120)
        else:
            print("\nNo main spine found for this hierarchy (no path starts with the first GA dimension or no matches).")
            print("Main spine subtree (full data structure): None")
        user_dims, virthiers = extract_user_defined_and_virthier(tree)
        print(f"User-defined dimensions: {', '.join(user_dims) if user_dims else 'None'}")
        print(f"Virtual hierarchy: {', '.join(virthiers) if virthiers else 'None'}")
        # Prepare Excel data
        # Level Comparison Table
        level_comparison, unmatched_ga, unmatched_cust = ordered_alignment_table(ga_list, main_spine)
        # Add extra unmatched customer dimensions as extra rows
        for dim in unmatched_cust:
            level_comparison.append(('', dim, '✗'))
        # Position Table
        position_table = []
        for i, dim in enumerate(main_spine):
            in_ga = dim.lower() in [g.lower() for g in ga_list]
            position_table.append((dim, in_ga))
        # Path Table
        path_table = []
        for path in all_paths:
            is_main = (path == main_spine)
            path_type = 'Main Spine' if is_main else 'Alternate Path'
            path_table.append((path_type, ' -> '.join(path), len(path), is_main))
        hierarchies_data[hier_name] = {
            'level_comparison': level_comparison,
            'user_dims': user_dims,
            'match_score': match_score,
            'virthiers': virthiers,
            'position_table': position_table,
            'path_table': path_table
        }
    # Print labeled intersections
    print("\nLabeled Intersections:")
    for k, v in labeled_intx.items():
        print(f"  {k}: {v}")
    # Write to Excel
    write_excel_output(hierarchies_data, labeled_intx)
    print("\nExcel output written to hierarchy_analysis.xlsx")

def get_file2_only_dimensions(xml_file: str, ga_file: str) -> dict:
    """
    Returns a dictionary:
    {
        'clnd': [...],
        'prod': [...],
        'loc': [...]
    }
    where each value is a list of dimension names in customer data that do not match with GA (file2_only for each hierarchy).
    """
    ga_hierarchies = read_ga_hierarchy_from_file(ga_file)
    hier_trees = extract_hierarchy_trees(xml_file)
    result = {'clnd': [], 'prod': [], 'loc': []}
    for hier_name, tree in hier_trees.items():
        all_paths = get_all_paths(tree)
        ga_list = ga_hierarchies.get(hier_name, [])
        main_spine, _ = find_main_spine(all_paths, ga_list)
        # file2_only: in customer (main spine) but not in GA
        file2_only = [dim for dim in main_spine if dim.lower() not in [g.lower() for g in ga_list]]
        level_comparison, unmatched_ga, unmatched_cust = ordered_alignment_table(ga_list, main_spine)
        result[hier_name] = unmatched_cust
    return result

def get_mismatched_dimensions(xml_file: str, ga_file: str) -> dict:
    """
    Returns a dictionary:
    {
        'clnd': {'ga_only': [...], 'file2_only': [...]},
        'prod': {'ga_only': [...], 'file2_only': [...]},
        'loc': {'ga_only': [...], 'file2_only': [...]}
    }
    where 'ga_only' are dimensions in GA but not in customer (main spine),
    and 'file2_only' are dimensions in customer (main spine) but not in GA.
    """
    ga_hierarchies = read_ga_hierarchy_from_file(ga_file)
    hier_trees = extract_hierarchy_trees(xml_file)
    result = {'clnd': {'ga_only': [], 'file2_only': []},
              'prod': {'ga_only': [], 'file2_only': []},
              'loc': {'ga_only': [], 'file2_only': []}}
    for hier_name, tree in hier_trees.items():
        all_paths = get_all_paths(tree)
        ga_list = ga_hierarchies.get(hier_name, [])
        main_spine, _ = find_main_spine(all_paths, ga_list)
        # file2_only: in customer (main spine) but not in GA
        file2_only = [dim for dim in main_spine if dim.lower() not in [g.lower() for g in ga_list]]
        # ga_only: in GA but not in customer (main spine)
        ga_only = [dim for dim in ga_list if dim.lower() not in [c.lower() for c in main_spine]]
        level_comparison, unmatched_ga, unmatched_cust = ordered_alignment_table(ga_list, main_spine)
        result[hier_name] = {'ga_only': unmatched_ga, 'file2_only': unmatched_cust}
    return result

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    ga_hierarchies = read_ga_hierarchy_from_file(GA_FILE)
    output_hierarchy_analysis(XML_FILE, ga_hierarchies)
    # Print file2_only dimensions for each hierarchy
    file2_only = get_file2_only_dimensions(XML_FILE, GA_FILE)
    print("\nfile2_only dimensions for each hierarchy:")
    for hier, dims in file2_only.items():
        print(f"  {hier}: {dims}")
    # Print mismatched dimensions for each hierarchy
    mismatched_dims = get_mismatched_dimensions(XML_FILE, GA_FILE)
    print("\nmismatched dimensions for each hierarchy:")
    for hier, dims in mismatched_dims.items():
        print(f"  {hier}:")
        print(f"    ga_only: {dims['ga_only']}")
        print(f"    file2_only: {dims['file2_only']}") 